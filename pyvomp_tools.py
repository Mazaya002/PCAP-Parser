import csv
from openpyxl import Workbook
import argparse
import os

# Function to read CSV file and convert it to a dictionary
def read_csv_to_dict(file_path):
    packets = {}
    with open(file_path, mode='r', encoding='utf-8') as file:
        reader = csv.DictReader(file)
        for row in reader:
            packets[row['tcp_payload']] = row  # Use tcp_payload as the key
    return packets

# Function to write data to an Excel sheet
def write_to_excel(data, sheet_name, workbook):
    sheet = workbook.create_sheet(title=sheet_name)
    
    if data:
        # Write header based on the keys of the first dictionary
        headers = list(data[0].keys())
        sheet.append(headers)

        # Write rows
        for row in data:
            sheet.append(list(row.values()))
    else:
        # If there is no data, just write a message in the sheet
        sheet.append(["No data found"])

# Function to compare packets between two CSV files and write results to Excel
def compare_packets(file1, file2, output_file):
    packets_file1 = read_csv_to_dict(file1)
    packets_file2 = read_csv_to_dict(file2)

    # Find matching, unique to file1, and unique to file2 packets
    matching_packets = [packets_file1[payload] for payload in packets_file1 if payload in packets_file2]
    unique_to_file1 = [packets_file1[payload] for payload in packets_file1 if payload not in packets_file2]
    unique_to_file2 = [packets_file2[payload] for payload in packets_file2 if payload not in packets_file1]

    # Create a new Excel workbook
    workbook = Workbook()

    # Write results to different sheets
    write_to_excel(matching_packets, 'Matching Packets', workbook)
    write_to_excel(unique_to_file1, 'Unique to File 1', workbook)
    write_to_excel(unique_to_file2, 'Unique to File 2', workbook)

    # Remove the default empty sheet created by openpyxl
    if 'Sheet' in workbook.sheetnames:
        std = workbook['Sheet']
        workbook.remove(std)

    # Save the Excel file
    workbook.save(output_file)
    print(f"Results written to {output_file}")

# Function to check if a file exists and has a .csv extension
def check_file_validity(file_path):
    if not os.path.isfile(file_path):
        raise FileNotFoundError(f"File '{file_path}' does not exist.")
    if not file_path.lower().endswith('.csv'):
        raise ValueError(f"File '{file_path}' is not a CSV file.")

# Main function to handle command-line arguments
def main():
    parser = argparse.ArgumentParser(description="Compare two CSV files and write the results to an Excel file.")
    
    # Arguments for file1, file2, and output file
    parser.add_argument('file1', help="Path to the first CSV file")
    parser.add_argument('file2', help="Path to the second CSV file")
    parser.add_argument('output', help="Path to the output Excel file")

    # Parse the arguments
    args = parser.parse_args()

    try:
        # Check the validity of the input CSV files
            check_file_validity(args.file1)
            check_file_validity(args.file2)

            # Check if the output file has a valid extension (optional)
            if not args.output.lower().endswith('.xlsx'):
                raise ValueError(f"Output file '{args.output}' must have an .xlsx extension.")

            # Call the compare_packets function with the arguments
            compare_packets(args.file1, args.file2, args.output)

    except (FileNotFoundError, ValueError) as e:
        print(f"Error: {e}")


if __name__ == "__main__":
    main()