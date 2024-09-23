# Python Parse comparer

# import csv  # Import csv module
# import pandas as pd

import csv
from openpyxl import Workbook

def read_csv_to_dict(file_path):
    packets = {}
    with open(file_path, mode='r', encoding='utf-8') as file:
        reader = csv.DictReader(file)
        for row in reader:
            packets[row['tcp_payload']] = row  # Use tcp_payload as the key
    return packets

def write_to_excel(data, sheet_name, workbook):
    sheet = workbook.create_sheet(title=sheet_name)
    
    if data:
        # Write header based on the keys of the first dictionary
        headers = list(data[0].keys())
        sheet.append(headers)

        # Write rows
        for row in data:
            sheet.append(list(row.values()))
    else:
        # If there is no data, just write a message in the sheet
        sheet.append(["No data found"])

def compare_packets(file1, file2, output_file):
    packets_file1 = read_csv_to_dict(file1)
    packets_file2 = read_csv_to_dict(file2)

    # Find matching, unique to file1, and unique to file2 packets
    matching_packets = [packets_file1[payload] for payload in packets_file1 if payload in packets_file2]
    unique_to_file1 = [packets_file1[payload] for payload in packets_file1 if payload not in packets_file2]
    unique_to_file2 = [packets_file2[payload] for payload in packets_file2 if payload not in packets_file1]

    # Create a new Excel workbook
    workbook = Workbook()

    # Write results to different sheets
    write_to_excel(matching_packets, 'Matching Packets', workbook)
    write_to_excel(unique_to_file1, 'Unique to File 1', workbook)
    write_to_excel(unique_to_file2, 'Unique to File 2', workbook)

    # Remove the default empty sheet created by openpyxl
    if 'Sheet' in workbook.sheetnames:
        std = workbook['Sheet']
        workbook.remove(std)

    # Save the Excel file
    workbook.save(output_file)
    print(f"Results written to {output_file}")

if __name__ == "__main__":
    compare_packets('Parse-file1.csv', 'Parse-file2.csv', 'comparison_results2.xlsx')







































# def read_csv_to_dict(file_path):
#     packets = {}
#     with open(file_path, mode='r', encoding='utf-8') as file:
#         reader = csv.DictReader(file)
#         for row in reader:
#             packets[row['tcp_payload']] = row  # Use payload as the key
#     return packets

# def compare_packets(file1, file2, output_file):
#     packets_file1 = read_csv_to_dict(file1)
#     packets_file2 = read_csv_to_dict(file2)

#     matching_packets = {payload: packets_file1[payload] for payload in packets_file1 if payload in packets_file2}
#     unique_to_file1 = {payload: packets_file1[payload] for payload in packets_file1 if payload not in packets_file2}
#     unique_to_file2 = {payload: packets_file2[payload] for payload in packets_file2 if payload not in packets_file1}

#     # Convert dict_values to list of dictionaries
#     matching_list = list(matching_packets.values())
#     unique_to_file1_list = list(unique_to_file1.values())
#     unique_to_file2_list = list(unique_to_file2.values())

#     # Create DataFrames
#     matching_df = pd.DataFrame(matching_list)
#     unique_to_file1_df = pd.DataFrame(unique_to_file1_list)
#     unique_to_file2_df = pd.DataFrame(unique_to_file2_list)

#     # Write to an Excel file with different sheets
#     with pd.ExcelWriter(output_file) as writer:
#         matching_df.to_excel(writer, sheet_name='Matching Packets', index=False)
#         unique_to_file1_df.to_excel(writer, sheet_name='Unique to File 1', index=False)
#         unique_to_file2_df.to_excel(writer, sheet_name='Unique to File 2', index=False)

#     print(f"Results written to {output_file}")

# if __name__ == "__main__":
#     compare_packets('Parse-file1.csv', 'Parse-file2.csv', 'comparison_results.xlsx')




    # compare_packets('Parse-file1.csv', 'Parse-file2.csv', 'comparison_results.xlsx')
